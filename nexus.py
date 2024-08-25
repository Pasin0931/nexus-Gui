import tkinter as tk
import openpyxl
import os
from tkinter import ttk


class ExcelApp():
    def __init__(self, root, path):
        self.frame = ttk.Frame(root)
        self.path = path
        self.combo_list = ["attendee", "staff", "speaker"]
        self.create_widgets()
        self.frame.pack()
        self.load_excel_data()

    def create_widgets(self):
        self.widgets_frame = ttk.LabelFrame(self.frame, text = "insert / edit")
        self.widgets_frame.grid(row=0, column=0, padx=20, pady=10)
        # entry 1
        self.name_entry = self.create_entry(self.widgets_frame, "Name", 0)
        self.age_spinbox = self.create_spinbox(self.widgets_frame, "Age", 1)
        self.role_combobox = self.create_combobox(self.widgets_frame, self.combo_list, 2)
        # entry 2
        self.email_entry = self.create_entry(self.widgets_frame, "Email", 3)

        self.button_frame = ttk.Frame(self.widgets_frame)
        self.button_frame.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

        self.create_button(self.button_frame, "Edit", self.edit_row, 0, 0)
        self.create_button(self.button_frame, "Delete", self.delete_row, 0, 1)
        self.create_button(self.widgets_frame, "Insert", self.insert_row, 5)

        self.create_treeview(self.frame)
        self.create_selected_frame(self.frame)

    def create_entry(self, parent, placeholder, row):
        entry = ttk.Entry(parent)
        entry.insert(0, placeholder)
        entry.bind("<FocusIn>", lambda e: entry.delete("0", "end"))
        entry.grid(row=row, column=0, padx=5, pady=(0, 5), sticky="ew")
        return entry
    
    def create_spinbox(self, parent, placeholder, row):
        spinbox = ttk.Spinbox(parent, from_=18, to=100)
        spinbox.insert(0, placeholder)
        spinbox.grid(row=row, column=0, padx=5, pady=5, sticky="ew")
        return spinbox
    
    def create_combobox(self, parent, values, row):
        combobox = ttk.Combobox(parent, values=values)
        combobox.current(0)
        combobox.grid(row=row, column=0, padx=5, pady=5, sticky="ew")
        return combobox
    
    def create_button(self, parent, text, command, row, column=0):
        button = ttk.Button(parent, text=text, command=command)
        button.grid(row=row, column=column, padx=5, pady=5, sticky="nsew")
    
    def edit_row(self):
        selected_item = self.treeview.selection()[0]

        name = self.name_entry.get()
        age = int(self.age_spinbox.get())
        role = self.role_combobox.get()
        email = self.email_entry.get()

        try:
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active
            
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[0].value == self.treeview.item(selected_item, "values")[0]:  # Match based on the Name
                    row[0].value = name
                    row[1].value = age
                    row[2].value = role
                    row[3].value = email
                    break

            workbook.save(self.path)
            self.treeview.item(selected_item, values=(name, age, role, email))
            self.clear_entries()
        except Exception as e:
            print(f"Error editing row: {e}")
    
    def delete_row(self):
        selected_item = self.treeview.selection()[0]
        name = self.treeview.item(selected_item, "values")[0]

        try:
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[0].value == name:  # Match based on the Name
                    sheet.delete_rows(row[0].row, 1)
                    break

            workbook.save(self.path)
            self.treeview.delete(selected_item)
            self.clear_entries()
        except Exception as e:
            print(f"Error deleting row: {e}")
    
    def insert_row(self):
        name = self.name_entry.get()
        age = self.age_spinbox.get()
        role = self.role_combobox.get()
        email = self.email_entry.get()

        row_val = [name, age, role, email]
        try:
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active
            sheet.append(row_val)
            
            workbook.save(self.path)
            self.treeview.insert('', tk.END, values=row_val)

            self.clear_entries()

        except Exception as a:
            print(f"Error inserting row: {a}")

        return print("insert")
    
    def create_treeview(self, parent):
        tree_frame = ttk.Frame(parent)
        tree_frame.grid(row=0, column=1, pady=10)
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side="right", fill="y")

        self.treeview = ttk.Treeview(tree_frame, show="headings",
                                     yscrollcommand=tree_scroll.set,
                                     columns=("Name", "Age", "Role", "Email"),
                                     height=13)
        self.treeview.column("Name", width=100)
        self.treeview.column("Age", width=50)
        self.treeview.column("Role", width=100)
        self.treeview.column("Email", width=200)
        self.treeview.pack()
        tree_scroll.config(command=self.treeview.yview)

        self.treeview.bind("<<TreeviewSelect>>", self.on_tree_select)

    def on_tree_select(self, event):
        selected_item = self.treeview.selection()[0]
        values = self.treeview.item(selected_item, "values")

        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, values[0])
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, values[1])
        self.role_combobox.set(values[2])
        self.email_entry.delete(0, "end")
        self.email_entry.insert(0, values[3])


        self.selected_name_label.config(text=values[0])
        self.selected_age_label.config(text=values[1])
        self.selected_role_label.config(text=values[2])
        self.selected_email_label.config(text=values[3])

    def clear_entries(self):
        self.name_entry.delete(0, "end")
        self.name_entry.insert(0, "Name")
        self.age_spinbox.delete(0, "end")
        self.age_spinbox.insert(0, "Age")
        self.role_combobox.set(self.combo_list[0])
        self.email_entry.delete(0, "end")
        self.email_entry.insert(0, "Email")

    def create_label(self, parent, text, row):
        ttk.Label(parent, text=text).grid(row=row, column=0, sticky="w")
        label = ttk.Label(parent, text="")
        label.grid(row=row, column=1, sticky="w")
        return label
    
    def create_selected_frame(self, parent):
        selected_frame = ttk.LabelFrame(parent, text="Information")
        selected_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=10, sticky="nsew")

        self.selected_name_label = self.create_label(selected_frame, "Name:", 0)
        self.selected_age_label = self.create_label(selected_frame, "Age:", 1)
        self.selected_role_label = self.create_label(selected_frame, "Role:", 2)
        self.selected_email_label = self.create_label(selected_frame, "Email:", 3)

    def load_excel_data(self):
        try:
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook.active

            list_val = list(sheet.values)
            for col_data in list_val[0]:
                self.treeview.heading(col_data, text=col_data)
            for value_tuple in list_val[1:]:
                self.treeview.insert("", tk.END, value=value_tuple)
            # print(list_val)
        except Exception as n:
            print(f"Error loading data: {n}")

if __name__ == "__main__":
    # path = os.path.join("./pep.xlsx")
    path = "D:/newSSgui/guiS/mainGui/pep.xlsx"
    root = tk.Tk()
    root.title("Example")
    app = ExcelApp(root, path)
    root.mainloop()